"""
test_parse_db_collections.py — parse_db collection rule + Name/Description
(EXPORT_PLAN §4). Builds an in-memory DB workbook with the REAL column names.

Rules under test:
  - a Ref used as another row's ParentRef is a "true collection" → filtered out
  - a leaf ET marked IsCollection='Y' (an app wrapper) is KEPT
  - Name and Description are read from their own real columns
  - ElementTypes keep _row_num (writable); PositionTypes strip it
"""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "backend"))

import openpyxl
import pytest
from parser import parse_db

ET_HEADERS = ["IsDeleted", "Ref", "IsCollection", "ParentRef", "Name",
              "Description", "SortOrder"]
PT_HEADERS = ["IsDeleted", "Ref", "IsCollection", "ParentRef", "Name",
              "Description", "DriverLocation"]

ET_ROWS = [
    # a parent category (used as ParentRef below) → true collection, filtered
    [None, "ET-CAT-DL", None, None, "Downlights", "Downlight category", 1],
    # leaf under that category → kept
    [None, "ET-DL-01", None, "ET-CAT-DL", "Downlight A", "A 3000K", 2],
    # app wrapper: IsCollection='Y' but NOT a parent of anything → KEPT
    [None, "ET-LIN-WRAP-01", "Y", None, "Linear wrapper", "LIN assembly", 3],
    # deleted → filtered
    ["Y", "ET-DEAD-01", None, None, "Dead", "gone", 4],
]

PT_ROWS = [
    [None, "PT-CAT", None, None, "Cat", "category", "Local"],
    [None, "PT-A01", None, "PT-CAT", "Position A", "desc A", "Local"],
]


def _make_db(tmp_dir):
    path = os.path.join(tmp_dir, "db.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    et = wb.create_sheet("ElementTypes")
    et.append(ET_HEADERS)
    for r in ET_ROWS:
        et.append(r)
    pt = wb.create_sheet("PositionTypes")
    pt.append(PT_HEADERS)
    for r in PT_ROWS:
        pt.append(r)
    wb.save(path)
    return path


def test_parent_ref_is_a_collection(tmp_path):
    data = parse_db(_make_db(str(tmp_path)))
    refs = {e["ElementTypeRef"] for e in data["element_types"]}
    assert "ET-CAT-DL" not in refs          # parent → collection, filtered
    assert "ET-DL-01" in refs               # leaf → kept


def test_wrapper_with_iscollection_flag_is_kept(tmp_path):
    data = parse_db(_make_db(str(tmp_path)))
    refs = {e["ElementTypeRef"] for e in data["element_types"]}
    # The landmine fix: IsCollection='Y' no longer causes filtering
    assert "ET-LIN-WRAP-01" in refs


def test_deleted_filtered(tmp_path):
    data = parse_db(_make_db(str(tmp_path)))
    refs = {e["ElementTypeRef"] for e in data["element_types"]}
    assert "ET-DEAD-01" not in refs


def test_name_and_description_are_separate(tmp_path):
    data = parse_db(_make_db(str(tmp_path)))
    dl = next(e for e in data["element_types"] if e["ElementTypeRef"] == "ET-DL-01")
    assert dl["Name"] == "Downlight A"
    assert dl["Description"] == "A 3000K"
    assert dl["Family"] == "ET-CAT-DL"       # ParentRef → Family


def test_element_types_keep_row_num_positions_do_not(tmp_path):
    data = parse_db(_make_db(str(tmp_path)))
    dl = next(e for e in data["element_types"] if e["ElementTypeRef"] == "ET-DL-01")
    assert "_row_num" in dl                   # writable
    pa = next(p for p in data["position_types"] if p["PositionTypeRef"] == "PT-A01")
    assert "_row_num" not in pa               # read-only
    assert pa["Name"] == "Position A"
    assert pa["Description"] == "desc A"
