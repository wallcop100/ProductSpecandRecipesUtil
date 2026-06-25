"""
test_patcher.py — Tests for backend/patcher.py

Run with:  pytest tests/backend/test_patcher.py  (from project root)
"""

import os
import re
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "backend"))

import openpyxl
import pytest
from patcher import backup_file, patch_ps, patch_rs

# ---------------------------------------------------------------------------
# Fixture paths
# ---------------------------------------------------------------------------

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "..", "fixtures")
PS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_ps.xlsx")
RS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_rs.xlsx")


def _require_fixture(path: str):
    if not os.path.isfile(path):
        pytest.skip(
            f"Fixture not found: {path}. "
            "Run 'python tests/fixtures/create_fixtures.py' to create it."
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _copy_fixture_to_tmp(src: str, tmp_dir: str) -> str:
    """Copy *src* to *tmp_dir* and return the new path."""
    dest = os.path.join(tmp_dir, os.path.basename(src))
    shutil.copy2(src, dest)
    return dest


# ---------------------------------------------------------------------------
# backup_file tests
# ---------------------------------------------------------------------------

class TestBackupFile:
    def test_backup_created(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        backup_path = backup_file(src)
        assert os.path.isfile(backup_path), f"Backup file not found: {backup_path}"

    def test_backup_in_same_directory(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        backup_path = backup_file(src)
        assert os.path.dirname(os.path.abspath(backup_path)) == str(tmp_path)

    def test_backup_filename_pattern(self, tmp_path):
        """Backup name must match <stem>.backup.YYYYMMDD_HHMMSS.xlsx"""
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        backup_path = backup_file(src)
        basename = os.path.basename(backup_path)
        pattern = r"^.+\.backup\.\d{8}_\d{6}\.xlsx$"
        assert re.match(pattern, basename), (
            f"Backup filename '{basename}' does not match expected pattern."
        )

    def test_backup_is_valid_xlsx(self, tmp_path):
        """The backup file must be a valid readable xlsx."""
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        backup_path = backup_file(src)
        wb = openpyxl.load_workbook(backup_path)
        assert "Form" in wb.sheetnames

    def test_original_unchanged_after_backup(self, tmp_path):
        """backup_file must not mutate the original file."""
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        original_size = os.path.getsize(src)
        backup_file(src)
        assert os.path.getsize(src) == original_size

    def test_backup_file_missing_source_raises(self, tmp_path):
        """backup_file should raise OSError when the source file doesn't exist."""
        missing = str(tmp_path / "nonexistent.xlsx")
        with pytest.raises(OSError):
            backup_file(missing)


# ---------------------------------------------------------------------------
# patch_ps tests
# ---------------------------------------------------------------------------

class TestPatchPs:
    def test_success_returns_true(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        result = patch_ps(src, [{"row_index": 2, "column": 2, "value": "NEW-CODE"}])
        assert result["success"] is True

    def test_backup_path_returned(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        result = patch_ps(src, [{"row_index": 2, "column": 2, "value": "NEW-CODE"}])
        assert "backup_path" in result
        assert os.path.isfile(result["backup_path"])

    def test_cell_value_updated(self, tmp_path):
        """After patch, the target cell in the saved file must have the new value."""
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        # Row 2, column 2 = ProductCode of first data row
        patch_ps(src, [{"row_index": 2, "column": 2, "value": "PATCHED-VALUE"}])
        wb = openpyxl.load_workbook(src)
        ws = wb["Form"]
        assert ws.cell(row=2, column=2).value == "PATCHED-VALUE"

    def test_multiple_changes_applied(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        changes = [
            {"row_index": 2, "column": 2, "value": "CODE-A"},
            {"row_index": 3, "column": 2, "value": "CODE-B"},
        ]
        patch_ps(src, changes)
        wb = openpyxl.load_workbook(src)
        ws = wb["Form"]
        assert ws.cell(row=2, column=2).value == "CODE-A"
        assert ws.cell(row=3, column=2).value == "CODE-B"

    def test_clear_cell_with_none(self, tmp_path):
        """Setting value=None should clear the cell."""
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        patch_ps(src, [{"row_index": 2, "column": 2, "value": None}])
        wb = openpyxl.load_workbook(src)
        ws = wb["Form"]
        assert ws.cell(row=2, column=2).value is None

    def test_missing_file_returns_error(self, tmp_path):
        missing = str(tmp_path / "missing_ps.xlsx")
        result = patch_ps(missing, [{"row_index": 2, "column": 2, "value": "X"}])
        assert result["success"] is False
        assert "error" in result

    def test_empty_changes_returns_error(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        result = patch_ps(src, [])
        assert result["success"] is False
        assert "error" in result

    def test_missing_row_index_key_returns_error(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        result = patch_ps(src, [{"column": 2, "value": "X"}])  # no row_index
        assert result["success"] is False

    def test_invalid_row_index_returns_error(self, tmp_path):
        _require_fixture(PS_FIXTURE)
        src = _copy_fixture_to_tmp(PS_FIXTURE, str(tmp_path))
        result = patch_ps(src, [{"row_index": 0, "column": 2, "value": "X"}])
        assert result["success"] is False


# ---------------------------------------------------------------------------
# patch_rs tests
# ---------------------------------------------------------------------------

class TestPatchRs:
    def test_success_returns_true(self, tmp_path):
        _require_fixture(RS_FIXTURE)
        src = _copy_fixture_to_tmp(RS_FIXTURE, str(tmp_path))
        result = patch_rs(src, [{"row_index": 2, "column": 10, "value": "Y"}])
        assert result["success"] is True

    def test_backup_path_returned(self, tmp_path):
        _require_fixture(RS_FIXTURE)
        src = _copy_fixture_to_tmp(RS_FIXTURE, str(tmp_path))
        result = patch_rs(src, [{"row_index": 2, "column": 10, "value": "Y"}])
        assert "backup_path" in result
        assert os.path.isfile(result["backup_path"])

    def test_cell_value_updated(self, tmp_path):
        """After patch, the target cell must reflect the new value."""
        _require_fixture(RS_FIXTURE)
        src = _copy_fixture_to_tmp(RS_FIXTURE, str(tmp_path))
        # Column 10 = IsDesign (1-based in RS Form sheet)
        patch_rs(src, [{"row_index": 2, "column": 10, "value": "Y"}])
        wb = openpyxl.load_workbook(src)
        ws = wb["Form"]
        assert ws.cell(row=2, column=10).value == "Y"

    def test_multiple_changes_applied(self, tmp_path):
        _require_fixture(RS_FIXTURE)
        src = _copy_fixture_to_tmp(RS_FIXTURE, str(tmp_path))
        changes = [
            {"row_index": 2, "column": 6, "value": 5},
            {"row_index": 3, "column": 6, "value": 10},
        ]
        patch_rs(src, changes)
        wb = openpyxl.load_workbook(src)
        ws = wb["Form"]
        assert ws.cell(row=2, column=6).value == 5
        assert ws.cell(row=3, column=6).value == 10

    def test_missing_file_returns_error(self, tmp_path):
        missing = str(tmp_path / "missing_rs.xlsx")
        result = patch_rs(missing, [{"row_index": 2, "column": 10, "value": "Y"}])
        assert result["success"] is False
        assert "error" in result

    def test_backup_unchanged_content(self, tmp_path):
        """The backup should contain the original value before patching."""
        _require_fixture(RS_FIXTURE)
        src = _copy_fixture_to_tmp(RS_FIXTURE, str(tmp_path))

        # Read original cell value
        wb_orig = openpyxl.load_workbook(src)
        original_value = wb_orig["Form"].cell(row=2, column=6).value
        wb_orig.close()

        result = patch_rs(src, [{"row_index": 2, "column": 6, "value": 999}])
        assert result["success"] is True

        # Backup must still hold the original value
        wb_backup = openpyxl.load_workbook(result["backup_path"])
        backup_value = wb_backup["Form"].cell(row=2, column=6).value
        wb_backup.close()

        assert backup_value == original_value
