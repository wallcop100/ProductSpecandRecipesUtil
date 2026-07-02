"""
test_patch_engine.py — Tests for app._apply_row_level_patch (the live patch
engine behind POST /patch). Covers the EXPORT_PLAN.md behaviours:

  - PS appends write EntityType='ElementType'
  - appends return assignments (reconciliation payload)
  - RS natural-key guard prevents duplicate appends
  - staleness check blocks the whole file and returns conflicts
  - deletes are tombstones (IsDeleted='Y'), never row removal
  - field-level RS patches touch only changed fields

Run with:  pytest tests/backend/test_patch_engine.py
"""

import os
import shutil
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "backend"))

import openpyxl
import pytest
from app import _apply_row_level_patch, PS_FIELD_TO_EXCEL, RS_FIELD_TO_EXCEL

# Real-format headers (from samplefiles/ — the canonical output format)
PS_HEADERS = [
    "IsDeleted", "EntityType", "EntityRef", "Manufacturer", "ProductCode",
    "CutPoint", "ComponentDescription", "InternalNotesText",
    "ExternalNotesText", "ComponentID", "CustomisationText",
    "ExplodeDescription", "ProductDescription", "IsTBC", "IsPropertiesTBC",
]
RS_HEADERS = [
    "ContextType", "ContextRef", "RecipeIndex", "EntityType", "EntityRef",
    "Sort Order", "RefSuffix", "Name", "Description", "Details", "Quantity",
    "PackQuantity", "IsDeleted", "_Notes", "IsDesign", "IsContractItem",
    "IsTRItem", "Dim_QuantityMultiplier", "IsInteger",
]

PS_ROWS = [
    # (EntityType, EntityRef, Manufacturer, ProductCode)
    ("ElementType", "ET-EXIST-01", "WAGO", "770-1112"),
    ("ElementType", "ET-EXIST-02", "Osram", "OS-500"),
]
RS_ROWS = [
    # (ContextType, ContextRef, RecipeIndex, EntityType, EntityRef, Quantity)
    ("PositionType", "PT-A01", 1, "ElementType", "ET-EXIST-01", 1),
    ("ElementType", "ET-LIN-01", 1, "ElementType", "ET-EXIST-02", 2),
]


def _make_ps(tmp_dir):
    path = os.path.join(tmp_dir, "ps.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"
    ws.append(PS_HEADERS)
    for et_type, ref, mfr, code in PS_ROWS:
        row = [None] * len(PS_HEADERS)
        row[PS_HEADERS.index("EntityType")] = et_type
        row[PS_HEADERS.index("EntityRef")] = ref
        row[PS_HEADERS.index("Manufacturer")] = mfr
        row[PS_HEADERS.index("ProductCode")] = code
        ws.append(row)
    wb.save(path)
    return path


def _make_rs(tmp_dir):
    path = os.path.join(tmp_dir, "rs.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"
    ws.append(RS_HEADERS)
    for ctx_type, ctx_ref, idx, et_type, ref, qty in RS_ROWS:
        row = [None] * len(RS_HEADERS)
        row[RS_HEADERS.index("ContextType")] = ctx_type
        row[RS_HEADERS.index("ContextRef")] = ctx_ref
        row[RS_HEADERS.index("RecipeIndex")] = idx
        row[RS_HEADERS.index("EntityType")] = et_type
        row[RS_HEADERS.index("EntityRef")] = ref
        row[RS_HEADERS.index("Quantity")] = qty
        ws.append(row)
    wb.save(path)
    return path


def _headers(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Form"]
    hdr = {str(c.value).strip(): i + 1
           for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
           if c.value is not None}
    wb.close()
    return hdr


def _cell(path, row, col_name):
    hdr = _headers(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    val = wb["Form"].cell(row=row, column=hdr[col_name]).value
    wb.close()
    return val


def _max_row(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    n = wb["Form"].max_row
    wb.close()
    return n


# ---------------------------------------------------------------------------
# PS
# ---------------------------------------------------------------------------

class TestPSAppend:
    def test_new_row_gets_entity_type(self, tmp_path):
        src = _make_ps(str(tmp_path))
        result = _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": "ET-NEW-01", "_isNew": True,
             "updates": {"Manufacturer": "Acme", "ProductCode": "A-100"}},
        ])
        assert result["success"]
        rn = result["assignments"]["ET-NEW-01"]
        assert _cell(src, rn, "EntityRef") == "ET-NEW-01"
        assert _cell(src, rn, "EntityType") == "ElementType"
        assert _cell(src, rn, "Manufacturer") == "Acme"

    def test_assignments_returned_for_appends_only(self, tmp_path):
        src = _make_ps(str(tmp_path))
        existing_ref = _cell(src, 2, "EntityRef")
        result = _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": existing_ref, "updates": {"Manufacturer": "Changed"}},
            {"elementTypeRef": "ET-NEW-02", "_isNew": True, "updates": {"ProductCode": "B-2"}},
        ])
        assert result["success"]
        assert existing_ref not in result["assignments"]
        assert "ET-NEW-02" in result["assignments"]

    def test_repeat_export_updates_not_duplicates(self, tmp_path):
        """Second export for the same new ref must update, not append again."""
        src = _make_ps(str(tmp_path))
        _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": "ET-NEW-03", "_isNew": True, "updates": {"ProductCode": "V1"}},
        ])
        rows_after_first = _max_row(src)
        result = _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": "ET-NEW-03", "_isNew": True, "updates": {"ProductCode": "V2"}},
        ])
        assert result["success"]
        assert _max_row(src) == rows_after_first  # no new row
        # value updated in place
        rn = rows_after_first
        assert _cell(src, rn, "ProductCode") == "V2"


class TestPSStaleness:
    def test_conflict_blocks_whole_file(self, tmp_path):
        src = _make_ps(str(tmp_path))
        ref = _cell(src, 2, "EntityRef")
        disk_mfr = _cell(src, 2, "Manufacturer")
        result = _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": ref,
             "updates": {"Manufacturer": "Mine"},
             "before": {"Manufacturer": "SomethingElse"}},   # stale base
        ])
        assert result["success"] is False
        assert result.get("conflict") is True
        assert len(result["conflicts"]) == 1
        c = result["conflicts"][0]
        assert c["key"] == ref and c["field"] == "Manufacturer"
        # nothing was written
        assert _cell(src, 2, "Manufacturer") == disk_mfr

    def test_matching_before_passes(self, tmp_path):
        src = _make_ps(str(tmp_path))
        ref = _cell(src, 2, "EntityRef")
        disk_mfr = _cell(src, 2, "Manufacturer")
        result = _apply_row_level_patch(src, "Form", PS_FIELD_TO_EXCEL, [
            {"elementTypeRef": ref,
             "updates": {"Manufacturer": "Mine"},
             "before": {"Manufacturer": disk_mfr}},
        ])
        assert result["success"]
        assert _cell(src, 2, "Manufacturer") == "Mine"


# ---------------------------------------------------------------------------
# RS
# ---------------------------------------------------------------------------

def _rs_row(**over):
    row = {
        "ContextType": "PositionType", "ContextRef": "PT-TEST",
        "RecipeIndex": 99, "ElementTypeRef": "ET-TEST-01",
        "Quantity": 1, "_row_num": None,
    }
    row.update(over)
    return row


class TestRSAppend:
    def test_append_returns_assignment_and_entity_type(self, tmp_path):
        src = _make_rs(str(tmp_path))
        result = _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "c1", "action": "upsert", "row": _rs_row()},
        ])
        assert result["success"]
        rn = result["assignments"]["c1"]
        assert _cell(src, rn, "EntityRef") == "ET-TEST-01"
        assert _cell(src, rn, "EntityType") == "ElementType"

    def test_natural_key_guard_prevents_duplicate(self, tmp_path):
        """Re-exporting an un-reconciled append must update, not duplicate."""
        src = _make_rs(str(tmp_path))
        _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "c1", "action": "upsert", "row": _rs_row(Quantity=1)},
        ])
        rows_after_first = _max_row(src)
        result = _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "c2", "action": "upsert", "row": _rs_row(Quantity=5)},
        ])
        assert result["success"]
        assert _max_row(src) == rows_after_first  # guarded — no duplicate
        rn = result["assignments"]["c2"]
        assert _cell(src, rn, "Quantity") == 5

    def test_delete_is_tombstone(self, tmp_path):
        src = _make_rs(str(tmp_path))
        rows_before = _max_row(src)
        result = _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "d1", "action": "delete", "row": {"_row_num": 2}},
        ])
        assert result["success"]
        assert _max_row(src) == rows_before          # row not removed
        assert _cell(src, 2, "IsDeleted") == "Y"     # tombstoned


class TestRSFieldLevel:
    def test_changed_fields_only(self, tmp_path):
        """Field-level patch must not touch other mapped columns."""
        src = _make_rs(str(tmp_path))
        original_ref = _cell(src, 2, "EntityRef")
        result = _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "f1", "action": "upsert",
             "row": {"_row_num": 2, "ElementTypeRef": "WRONG-STALE-VALUE"},
             "changedFields": {"Quantity": 42}},
        ])
        assert result["success"]
        assert _cell(src, 2, "Quantity") == 42
        assert _cell(src, 2, "EntityRef") == original_ref  # untouched

    def test_staleness_conflict_blocks(self, tmp_path):
        src = _make_rs(str(tmp_path))
        disk_qty = _cell(src, 2, "Quantity")
        result = _apply_row_level_patch(src, "Form", RS_FIELD_TO_EXCEL, [
            {"_id": "f1", "action": "upsert",
             "row": {"_row_num": 2},
             "changedFields": {"Quantity": 42},
             "before": {"Quantity": 12345}},   # stale base
        ])
        assert result["success"] is False
        assert result.get("conflict") is True
        assert _cell(src, 2, "Quantity") == disk_qty  # untouched


# ---------------------------------------------------------------------------
# DB ElementTypes (writable catalogue — EXPORT_PLAN §4)
# ---------------------------------------------------------------------------

from app import DB_FIELD_TO_EXCEL

DB_ET_HEADERS = ["IsDeleted", "Ref", "IsCollection", "ParentRef", "Name",
                 "Description", "SortOrder"]
DB_ET_ROWS = [
    [None, "ET-EXIST-01", None, "ET-CAT", "Existing A", "desc A", 2],
    [None, "ET-EXIST-02", None, "ET-CAT", "Existing B", "desc B", 3],
]


def _make_db_wb(tmp_dir):
    path = os.path.join(tmp_dir, "db.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    et = wb.create_sheet("ElementTypes")
    et.append(DB_ET_HEADERS)
    for r in DB_ET_ROWS:
        et.append(r)
    wb.save(path)
    return path


def _db_cell(path, row, col_name):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["ElementTypes"]
    hdr = {str(c.value).strip(): i + 1
           for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
           if c.value is not None}
    val = ws.cell(row=row, column=hdr[col_name]).value
    wb.close()
    return val


class TestDBCatalogue:
    def test_append_new_element_type(self, tmp_path):
        src = _make_db_wb(str(tmp_path))
        result = _apply_row_level_patch(
            src, "ElementTypes", DB_FIELD_TO_EXCEL,
            [{"elementTypeRef": "ET-NEW-01", "_isNew": True,
              "updates": {"Name": "New One", "Description": "brand new",
                          "Family": "ET-CAT", "SortOrder": 4}}],
            key_column="Ref")
        assert result["success"]
        rn = result["assignments"]["ET-NEW-01"]
        assert _db_cell(src, rn, "Ref") == "ET-NEW-01"
        assert _db_cell(src, rn, "Name") == "New One"
        assert _db_cell(src, rn, "Description") == "brand new"
        assert _db_cell(src, rn, "ParentRef") == "ET-CAT"

    def test_rename_updates_ref_in_place(self, tmp_path):
        """Rename = look up by old ref, write the new ref into the Ref cell."""
        src = _make_db_wb(str(tmp_path))
        result = _apply_row_level_patch(
            src, "ElementTypes", DB_FIELD_TO_EXCEL,
            [{"elementTypeRef": "ET-EXIST-01",
              "updates": {"ElementTypeRef": "ET-RENAMED-01"}}],
            key_column="Ref")
        assert result["success"]
        assert _db_cell(src, 2, "Ref") == "ET-RENAMED-01"

    def test_soft_delete(self, tmp_path):
        src = _make_db_wb(str(tmp_path))
        result = _apply_row_level_patch(
            src, "ElementTypes", DB_FIELD_TO_EXCEL,
            [{"elementTypeRef": "ET-EXIST-02", "updates": {"IsDeleted": "Y"}}],
            key_column="Ref")
        assert result["success"]
        assert _db_cell(src, 3, "IsDeleted") == "Y"
        assert _db_cell(src, 3, "Ref") == "ET-EXIST-02"   # row not removed

    def test_staleness_blocks(self, tmp_path):
        src = _make_db_wb(str(tmp_path))
        result = _apply_row_level_patch(
            src, "ElementTypes", DB_FIELD_TO_EXCEL,
            [{"elementTypeRef": "ET-EXIST-01",
              "updates": {"Name": "Mine"}, "before": {"Name": "STALE"}}],
            key_column="Ref")
        assert result["success"] is False
        assert result.get("conflict") is True
        assert _db_cell(src, 2, "Name") == "Existing A"   # untouched
