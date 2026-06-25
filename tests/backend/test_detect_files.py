"""
test_detect_files.py — Tests for the detect_files / detect_file_type logic in backend/parser.py

Run with:  pytest tests/backend/test_detect_files.py  (from project root)
"""

import os
import shutil
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "backend"))

import openpyxl
import pytest
from parser import detect_file_type, detect_files

# ---------------------------------------------------------------------------
# Fixture paths
# ---------------------------------------------------------------------------

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "..", "fixtures")
DB_FIXTURE = os.path.join(FIXTURE_DIR, "sample_db.xlsx")
PS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_ps.xlsx")
RS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_rs.xlsx")


def _require_fixture(path: str):
    if not os.path.isfile(path):
        pytest.skip(
            f"Fixture not found: {path}. "
            "Run 'python tests/fixtures/create_fixtures.py' to create it."
        )


# ---------------------------------------------------------------------------
# Helpers to build minimal test xlsx files
# ---------------------------------------------------------------------------

def _make_db_xlsx(path: str) -> str:
    """Write a minimal DB-type xlsx (has ElementTypes + PositionTypes sheets)."""
    wb = openpyxl.Workbook()
    wb.active.title = "ElementTypes"
    wb.create_sheet("PositionTypes")
    wb.save(path)
    return path


def _make_ps_xlsx(path: str) -> str:
    """Write a minimal PS-type xlsx (has Form sheet with ProductCode header)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"
    ws.append(["ElementTypeRef", "ProductCode", "SupplierCode"])
    wb.save(path)
    return path


def _make_rs_xlsx(path: str) -> str:
    """Write a minimal RS-type xlsx (has Form sheet with ContextRef header)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"
    ws.append(["PositionTypeRef", "ContextRef", "ElementTypeRef"])
    wb.save(path)
    return path


def _make_unknown_xlsx(path: str) -> str:
    """Write an xlsx that doesn't match any known type."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Column1", "Column2"])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# detect_file_type tests
# ---------------------------------------------------------------------------

class TestDetectFileType:
    def test_identifies_db_fixture(self):
        _require_fixture(DB_FIXTURE)
        assert detect_file_type(DB_FIXTURE) == "db"

    def test_identifies_ps_fixture(self):
        _require_fixture(PS_FIXTURE)
        assert detect_file_type(PS_FIXTURE) == "ps"

    def test_identifies_rs_fixture(self):
        _require_fixture(RS_FIXTURE)
        assert detect_file_type(RS_FIXTURE) == "rs"

    def test_identifies_minimal_db(self, tmp_path):
        path = _make_db_xlsx(str(tmp_path / "db.xlsx"))
        assert detect_file_type(path) == "db"

    def test_identifies_minimal_ps(self, tmp_path):
        path = _make_ps_xlsx(str(tmp_path / "ps.xlsx"))
        assert detect_file_type(path) == "ps"

    def test_identifies_minimal_rs(self, tmp_path):
        path = _make_rs_xlsx(str(tmp_path / "rs.xlsx"))
        assert detect_file_type(path) == "rs"

    def test_unknown_file_returns_none(self, tmp_path):
        path = _make_unknown_xlsx(str(tmp_path / "unknown.xlsx"))
        assert detect_file_type(path) is None

    def test_non_existent_file_returns_none(self, tmp_path):
        missing = str(tmp_path / "nonexistent.xlsx")
        assert detect_file_type(missing) is None

    def test_non_xlsx_file_returns_none(self, tmp_path):
        """A .txt file should not be mistaken for any known type."""
        txt_file = str(tmp_path / "notanexcel.txt")
        with open(txt_file, "w") as f:
            f.write("some text\n")
        # detect_file_type tries to open as xlsx → should return None (not crash)
        assert detect_file_type(txt_file) is None

    def test_db_prioritised_over_form_sheet(self, tmp_path):
        """If a file has ElementTypes + PositionTypes AND a Form sheet, classify as DB."""
        wb = openpyxl.Workbook()
        wb.active.title = "ElementTypes"
        wb.create_sheet("PositionTypes")
        ws_form = wb.create_sheet("Form")
        ws_form.append(["ProductCode"])  # would match PS if checked
        path = str(tmp_path / "ambiguous.xlsx")
        wb.save(path)
        assert detect_file_type(path) == "db"


# ---------------------------------------------------------------------------
# detect_files tests
# ---------------------------------------------------------------------------

class TestDetectFiles:
    def test_returns_expected_structure(self, tmp_path):
        result = detect_files(str(tmp_path))
        assert "db" in result
        assert "ps" in result
        assert "rs" in result
        assert "all_xlsx" in result

    def test_empty_folder_returns_nulls(self, tmp_path):
        result = detect_files(str(tmp_path))
        assert result["db"] is None
        assert result["ps"] is None
        assert result["rs"] is None
        assert result["all_xlsx"] == []

    def test_detects_db_file(self, tmp_path):
        _make_db_xlsx(str(tmp_path / "mydb.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["db"] == "mydb.xlsx"

    def test_detects_ps_file(self, tmp_path):
        _make_ps_xlsx(str(tmp_path / "myps.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["ps"] == "myps.xlsx"

    def test_detects_rs_file(self, tmp_path):
        _make_rs_xlsx(str(tmp_path / "myrs.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["rs"] == "myrs.xlsx"

    def test_detects_all_three(self, tmp_path):
        _make_db_xlsx(str(tmp_path / "a_db.xlsx"))
        _make_ps_xlsx(str(tmp_path / "b_ps.xlsx"))
        _make_rs_xlsx(str(tmp_path / "c_rs.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["db"] == "a_db.xlsx"
        assert result["ps"] == "b_ps.xlsx"
        assert result["rs"] == "c_rs.xlsx"

    def test_all_xlsx_includes_all_files(self, tmp_path):
        _make_db_xlsx(str(tmp_path / "a_db.xlsx"))
        _make_ps_xlsx(str(tmp_path / "b_ps.xlsx"))
        _make_rs_xlsx(str(tmp_path / "c_rs.xlsx"))
        result = detect_files(str(tmp_path))
        assert set(result["all_xlsx"]) == {"a_db.xlsx", "b_ps.xlsx", "c_rs.xlsx"}

    def test_all_xlsx_is_sorted(self, tmp_path):
        _make_db_xlsx(str(tmp_path / "z_file.xlsx"))
        _make_ps_xlsx(str(tmp_path / "a_file.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["all_xlsx"] == sorted(result["all_xlsx"])

    def test_ignores_non_xlsx_files(self, tmp_path):
        txt_path = str(tmp_path / "readme.txt")
        with open(txt_path, "w") as f:
            f.write("not xlsx\n")
        csv_path = str(tmp_path / "data.csv")
        with open(csv_path, "w") as f:
            f.write("a,b,c\n")
        result = detect_files(str(tmp_path))
        # all_xlsx must not include .txt or .csv
        for name in result["all_xlsx"]:
            assert name.lower().endswith(".xlsx")

    def test_unknown_xlsx_in_all_xlsx_but_not_typed(self, tmp_path):
        _make_unknown_xlsx(str(tmp_path / "unknown.xlsx"))
        result = detect_files(str(tmp_path))
        assert "unknown.xlsx" in result["all_xlsx"]
        assert result["db"] is None
        assert result["ps"] is None
        assert result["rs"] is None

    def test_fixture_folder_detects_all_three(self):
        """Smoke test: the real fixture directory contains all three file types."""
        for p in (DB_FIXTURE, PS_FIXTURE, RS_FIXTURE):
            _require_fixture(p)
        result = detect_files(FIXTURE_DIR)
        assert result["db"] is not None
        assert result["ps"] is not None
        assert result["rs"] is not None

    def test_nonexistent_folder_returns_empty(self, tmp_path):
        missing_dir = str(tmp_path / "no_such_folder")
        result = detect_files(missing_dir)
        assert result["db"] is None
        assert result["ps"] is None
        assert result["rs"] is None
        assert result["all_xlsx"] == []

    def test_returns_filename_not_full_path(self, tmp_path):
        """detect_files should return just the filename, not the full path."""
        _make_ps_xlsx(str(tmp_path / "ps_file.xlsx"))
        result = detect_files(str(tmp_path))
        assert result["ps"] == "ps_file.xlsx"
        # Must not be a full path
        assert not os.path.isabs(result["ps"])
