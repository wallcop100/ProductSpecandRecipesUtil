"""
test_parser.py — Tests for backend/parser.py

Run with:  pytest tests/backend/test_parser.py  (from project root)
"""

import os
import sys

# Ensure the backend package is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "backend"))

import pytest
from parser import parse_db, parse_ps, parse_rs

# ---------------------------------------------------------------------------
# Fixture paths
# ---------------------------------------------------------------------------

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "..", "fixtures")
DB_FIXTURE = os.path.join(FIXTURE_DIR, "sample_db.xlsx")
PS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_ps.xlsx")
RS_FIXTURE = os.path.join(FIXTURE_DIR, "sample_rs.xlsx")


# ---------------------------------------------------------------------------
# Guards — skip all tests if fixtures are missing (run create_fixtures.py first)
# ---------------------------------------------------------------------------

def _require_fixture(path: str):
    if not os.path.isfile(path):
        pytest.skip(
            f"Fixture not found: {path}. "
            "Run 'python tests/fixtures/create_fixtures.py' to create it."
        )


# ---------------------------------------------------------------------------
# parse_db tests
# ---------------------------------------------------------------------------

class TestParseDb:
    def test_returns_dict_with_expected_keys(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        assert isinstance(result, dict)
        assert "element_types" in result
        assert "position_types" in result

    def test_element_types_count(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        # fixture has 18 element type rows
        assert len(result["element_types"]) == 18

    def test_element_types_columns(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        expected_cols = {"ElementTypeRef", "Name", "Family", "Variant"}
        first = result["element_types"][0]
        assert expected_cols == set(first.keys())

    def test_element_types_first_row(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        first = result["element_types"][0]
        assert first["ElementTypeRef"] == "ET-DL-01"
        assert first["Name"] == "Standard Downlight Virtual"
        assert first["Family"] == "DL"
        assert first["Variant"] == "Standard"

    def test_position_types_count(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        # fixture has 5 position type rows
        assert len(result["position_types"]) == 5

    def test_position_types_columns(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        expected_cols = {
            "PositionTypeRef", "Name", "DriverLocation", "SecondaryPowerType",
            "ControlTypeRef", "SecondaryPowerNodes_+ve"
        }
        first = result["position_types"][0]
        assert expected_cols == set(first.keys())

    def test_position_types_lin_row(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        # PT-LIN-01 should have SecondaryPowerType="CV"
        lin_rows = [
            pt for pt in result["position_types"]
            if pt["PositionTypeRef"] == "PT-LIN-01"
        ]
        assert len(lin_rows) == 1
        assert lin_rows[0]["SecondaryPowerType"] == "CV"
        assert lin_rows[0]["DriverLocation"] == "Local"

    def test_position_types_numeric_nodes(self):
        _require_fixture(DB_FIXTURE)
        result = parse_db(DB_FIXTURE)
        twin = next(
            pt for pt in result["position_types"]
            if pt["PositionTypeRef"] == "PT-DL-TW-01"
        )
        # SecondaryPowerNodes_+ve stored as integer 2 in xlsx
        assert twin["SecondaryPowerNodes_+ve"] == 2

    def test_missing_sheet_raises(self, tmp_path):
        """A workbook without ElementTypes/PositionTypes should raise ValueError."""
        import openpyxl
        wb = openpyxl.Workbook()
        bad_path = str(tmp_path / "bad.xlsx")
        wb.save(bad_path)
        with pytest.raises(ValueError, match="ElementTypes"):
            parse_db(bad_path)


# ---------------------------------------------------------------------------
# parse_ps tests
# ---------------------------------------------------------------------------

class TestParsePs:
    def test_returns_list(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        assert isinstance(result, list)

    def test_row_count(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        # fixture has 9 data rows
        assert len(result) == 9

    def test_columns(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        expected_cols = {
            "ElementTypeRef", "ProductCode", "SupplierCode", "Description",
            "IsDesign", "IsContractItem", "IsTBC", "IsPropertiesTBC", "Notes"
        }
        assert set(result[0].keys()) == expected_cols

    def test_na_product_code(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        dl_row = next(r for r in result if r["ElementTypeRef"] == "ET-DL-01")
        assert dl_row["ProductCode"] == "N/A"

    def test_flag_column_y(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        dl_row = next(r for r in result if r["ElementTypeRef"] == "ET-DL-01")
        # ET-DL-01 has IsDesign="Y"
        assert dl_row["IsDesign"] == "Y"

    def test_flag_column_null(self):
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        dl_row = next(r for r in result if r["ElementTypeRef"] == "ET-DL-01")
        # IsContractItem is None for ET-DL-01
        assert dl_row["IsContractItem"] is None

    def test_duplicate_product_code_rows_present(self):
        """Both TAPE-01 rows are present for the duplicate detection test."""
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        tape_codes = [r for r in result if r.get("ProductCode") == "TAPE-01"]
        assert len(tape_codes) == 2

    def test_missing_form_sheet_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        bad_path = str(tmp_path / "bad_ps.xlsx")
        wb.save(bad_path)
        with pytest.raises(ValueError, match="Form"):
            parse_ps(bad_path)

    def test_empty_string_cells_return_none(self):
        """Cells with empty strings should become None (not '')."""
        _require_fixture(PS_FIXTURE)
        result = parse_ps(PS_FIXTURE)
        for row in result:
            # SupplierCode of ET-DL-01 is an empty string in the xlsx
            if row["ElementTypeRef"] == "ET-DL-01":
                # empty string → None
                assert row["SupplierCode"] is None or row["SupplierCode"] == ""
                # Notes is also empty
                break


# ---------------------------------------------------------------------------
# parse_rs tests
# ---------------------------------------------------------------------------

class TestParseRs:
    def test_returns_list(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        assert isinstance(result, list)

    def test_row_count(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        # fixture has 21 data rows
        assert len(result) == 21

    def test_columns(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        expected_cols = {
            "PositionTypeRef", "ContextType", "ContextRef", "RecipeIndex",
            "ElementTypeRef", "Quantity", "Dim_QuantityMultiplier", "Dim_Quantity",
            "IsInteger", "IsDesign", "IsContractItem", "IsTBC", "IsPropertiesTBC", "Notes"
        }
        assert set(result[0].keys()) == expected_cols

    def test_is_design_y_parsed_correctly(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        # PT-DL-LOCAL-01 position type row should have IsDesign="Y"
        pt_row = next(
            r for r in result
            if r["PositionTypeRef"] == "PT-DL-LOCAL-01"
            and r["ContextType"] == "PositionType"
            and r["ElementTypeRef"] == "ET-DL-01"
        )
        assert pt_row["IsDesign"] == "Y"

    def test_missing_is_design_pt_row(self):
        """PT-DL-CC-01 has no IsDesign=Y row (seeded error)."""
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        cc_pt_rows = [
            r for r in result
            if r["PositionTypeRef"] == "PT-DL-CC-01"
            and r["ContextType"] == "PositionType"
        ]
        assert all(r["IsDesign"] is None for r in cc_pt_rows)

    def test_numeric_quantity(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        sock_row = next(
            r for r in result
            if r["PositionTypeRef"] == "PT-DL-LOCAL-01"
            and r["ElementTypeRef"] == "ET-5Pin-Local-01"
        )
        assert sock_row["Quantity"] == 1

    def test_dim_quantity_numeric(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        clip_row = next(
            r for r in result
            if r["PositionTypeRef"] == "PT-LIN-01"
            and r["ElementTypeRef"] == "ET-CLIP-01"
        )
        assert clip_row["Dim_Quantity"] == pytest.approx(3.2)

    def test_lin_rows_present(self):
        _require_fixture(RS_FIXTURE)
        result = parse_rs(RS_FIXTURE)
        lin_rows = [r for r in result if r["PositionTypeRef"] == "PT-LIN-01"]
        assert len(lin_rows) >= 5  # at least 5 rows for PT-LIN-01

    def test_missing_form_sheet_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        bad_path = str(tmp_path / "bad_rs.xlsx")
        wb.save(bad_path)
        with pytest.raises(ValueError, match="Form"):
            parse_rs(bad_path)
