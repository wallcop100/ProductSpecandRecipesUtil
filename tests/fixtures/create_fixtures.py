"""
create_fixtures.py — Generate xlsx fixture files for Recipe Builder tests.

Run from anywhere::

    python tests/fixtures/create_fixtures.py

Output files are written to the same directory as this script (tests/fixtures/).
"""

from __future__ import annotations

import os
import sys

# ---------------------------------------------------------------------------
# Bootstrap: ensure openpyxl is importable regardless of CWD
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("ERROR: openpyxl is not installed.  Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

FIXTURE_DIR = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, sheet_name: str, headers: list, rows: list) -> None:
    """Add *sheet_name* to *wb* and populate it with *headers* then *rows*."""
    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _save(wb: Workbook, filename: str) -> str:
    path = os.path.join(FIXTURE_DIR, filename)
    wb.save(path)
    print(f"  Created: {path}")
    return path


# ---------------------------------------------------------------------------
# sample_db.xlsx
# ---------------------------------------------------------------------------

def create_sample_db() -> str:
    wb = Workbook()
    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # --- ElementTypes sheet ---
    et_headers = ["ElementTypeRef", "Name", "Family", "Variant"]
    et_rows = [
        ["ET-DL-01",                "Standard Downlight Virtual",     "DL",      "Standard"],
        ["ET-CCL-D-700-1CH-01",     "Local DALI Driver 700mA",        "CCL",     "D-700-1CH"],
        ["ET-5Pin-Local-01",        "5-Pin Local Socket",             "5Pin",    "Local"],
        ["ET-5Pin-SR-01",           "5-Pin Strain Relief",            "5Pin",    "SR"],
        ["ET-2Pin-DC-Socket-01",    "2-Pin DC Socket",                "2Pin",    "DC-Socket"],
        ["ET-2Pin-DC-Plug-01",      "2-Pin DC Plug",                  "2Pin",    "DC-Plug"],
        ["ET-2Pin-DC-SR-01",        "2-Pin DC Strain Relief",         "2Pin",    "DC-SR"],
        ["ET-LIN-01",               "Linear Virtual Element",         "LIN",     "Standard"],
        ["ET-LIN-Socket-01",        "LIN 2-Pin Socket",               "LIN",     "Socket"],
        ["ET-LIN-Plug-01",          "LIN 2-Pin Plug",                 "LIN",     "Plug"],
        ["ET-LLOCK-01",             "Locking Lever",                  "LLOCK",   "Standard"],
        ["ET-CLIP-01",              "LIN Clip",                       "CLIP",    "Standard"],
        ["ET-TAPE-01",              "LED Tape",                       "TAPE",    "Standard"],
        ["ET-PROFILE-01",           "Aluminium Profile",              "PROFILE", "Standard"],
        ["ET-DIFF-01",              "Diffuser",                       "DIFF",    "Standard"],
        ["ET-CAP-01",               "End Cap",                        "CAP",     "Standard"],
        ["ET-2Pin-Remote-Socket-01","2-Pin Remote CC Socket",         "2Pin",    "Remote-Socket"],
        ["ET-2Pin-Remote-Plug-01",  "2-Pin Remote CC Plug",           "2Pin",    "Remote-Plug"],
    ]
    _write_sheet(wb, "ElementTypes", et_headers, et_rows)

    # --- PositionTypes sheet ---
    pt_headers = [
        "PositionTypeRef", "Name", "DriverLocation", "SecondaryPowerType",
        "ControlTypeRef", "SecondaryPowerNodes_+ve"
    ]
    pt_rows = [
        ["PT-DL-LOCAL-01", "Standard Local Downlight", "Local",  "CC", "DALI",  1],
        ["PT-DL-CC-01",    "Remote CC Downlight",       "Remote", "CC", "DALI",  1],
        ["PT-LIN-01",      "Linear Tape Profile",        "Local",  "CV", "LOCAL", 1],
        ["PT-DL-EXT-01",   "Exterior Downlight",         "Local",  "CC", "DALI",  1],
        ["PT-DL-TW-01",    "Twin Spot Downlight",        "Local",  "CC", "DALI",  2],
    ]
    _write_sheet(wb, "PositionTypes", pt_headers, pt_rows)

    return _save(wb, "sample_db.xlsx")


# ---------------------------------------------------------------------------
# sample_ps.xlsx
# ---------------------------------------------------------------------------

def create_sample_ps() -> str:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ps_headers = [
        "ElementTypeRef", "ProductCode", "SupplierCode", "Description",
        "IsDesign", "IsContractItem", "IsTBC", "IsPropertiesTBC", "Notes"
    ]
    # fmt: off
    # (None represents an empty / null cell)
    ps_rows = [
        # ElementTypeRef              ProductCode      SupplierCode  Description                    IsDesign  IsContractItem  IsTBC  IsPropertiesTBC  Notes
        ["ET-DL-01",                 "N/A",           "",           "Standard Downlight",          "Y",      None,           None,  None,            ""],
        ["ET-CCL-D-700-1CH-01",      "DRV-700-1CH",   "SUPPLIER-A", "700mA DALI Driver",           None,     "Y",            None,  None,            ""],
        ["ET-5Pin-Local-01",         "SOCK-5P-LOC",   "SUPPLIER-B", "5-Pin Local Socket",          None,     "Y",            None,  None,            ""],
        ["ET-5Pin-SR-01",            "SR-5P",         "SUPPLIER-B", "5-Pin Strain Relief",         None,     "Y",            None,  None,            ""],
        ["ET-2Pin-DC-Socket-01",     "SOCK-2P-DC",    "SUPPLIER-C", "2-Pin DC Socket",             None,     "Y",            None,  None,            ""],
        ["ET-2Pin-DC-Plug-01",       "PLUG-2P-DC",    "SUPPLIER-C", "2-Pin DC Plug",               None,     "Y",            None,  None,            ""],
        ["ET-LIN-01",                "N/A",           "",           "Linear Virtual Element",      "Y",      None,           None,  None,            ""],
        # TAPE-01 used for both ET-TAPE-01 and ET-PROFILE-01 → seeds DUPLICATE_PRODUCT_CODE error
        ["ET-TAPE-01",               "TAPE-01",       "SUPPLIER-D", "LED Tape",                    None,     None,           None,  None,            ""],
        ["ET-PROFILE-01",            "TAPE-01",       "SUPPLIER-D", "Profile",                     None,     None,           None,  None,            ""],
    ]
    # fmt: on
    _write_sheet(wb, "Form", ps_headers, ps_rows)

    return _save(wb, "sample_ps.xlsx")


# ---------------------------------------------------------------------------
# sample_rs.xlsx
# ---------------------------------------------------------------------------

def create_sample_rs() -> str:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    rs_headers = [
        "PositionTypeRef", "ContextType", "ContextRef", "RecipeIndex",
        "ElementTypeRef", "Quantity", "Dim_QuantityMultiplier", "Dim_Quantity",
        "IsInteger", "IsDesign", "IsContractItem", "IsTBC", "IsPropertiesTBC", "Notes"
    ]

    # fmt: off
    rs_rows = [
        # ─── PT-DL-LOCAL-01 — complete, valid recipe ───────────────────────────
        # PositionTypeRef       ContextType    ContextRef          Idx  ElementTypeRef           Qty   DimMult  DimQty  IsInt  IsDes  IsCI  IsTBC  IsPropTBC  Notes
        ["PT-DL-LOCAL-01", "PositionType", "PT-DL-LOCAL-01", 1, "ET-DL-01",               None, None,  None,   None,  "Y",   None,  None,  None,  ""],
        ["PT-DL-LOCAL-01", "PositionType", "PT-DL-LOCAL-01", 2, "ET-5Pin-Local-01",        1,    None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-LOCAL-01", "PositionType", "PT-DL-LOCAL-01", 3, "ET-5Pin-SR-01",           1,    None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-LOCAL-01", "ElementType",  "ET-DL-01",       1, "ET-CCL-D-700-1CH-01",    1,    None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-LOCAL-01", "ElementType",  "ET-DL-01",       2, "ET-2Pin-DC-Socket-01",   1,    None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-LOCAL-01", "ElementType",  "ET-DL-01",       3, "ET-2Pin-DC-Plug-01",     1,    None,  None,   None,  None,  "Y",   None,  None,  ""],

        # ─── PT-DL-CC-01 — MISSING IsDesign (seeded error) ────────────────────
        # No row has IsDesign="Y" for this position type
        ["PT-DL-CC-01",    "PositionType", "PT-DL-CC-01",    1, "ET-DL-01",               None, None,  None,   None,  None,  None,  None,  None,  ""],
        ["PT-DL-CC-01",    "ElementType",  "ET-DL-01",       1, "ET-2Pin-Remote-Socket-01",1,   None,  None,   None,  None,  "Y",   None,  None,  ""],

        # ─── PT-LIN-01 — MISSING locking lever (seeded error) ─────────────────
        # Also seeds MISSING_CLIPS_DIM_QTY warning (Dim_Quantity is null for CLIP)
        ["PT-LIN-01",      "PositionType", "PT-LIN-01",      1, "ET-LIN-01",              None, None,  None,   None,  "Y",   None,  None,  None,  ""],
        ["PT-LIN-01",      "PositionType", "PT-LIN-01",      2, "ET-LIN-Socket-01",        1,   None,  None,   None,  None,  "Y",   None,  None,  ""],
        # CLIP row with a Dim_Quantity set (so this clip row is OK)
        ["PT-LIN-01",      "PositionType", "PT-LIN-01",      3, "ET-CLIP-01",             None, None,  3.2,    "Y",   None,  "Y",   None,  None,  ""],
        # No ET-LLOCK-01 row → MISSING_LOCKING_LEVER fires
        ["PT-LIN-01",      "ElementType",  "ET-LIN-01",      1, "ET-TAPE-01",             None, 1,     None,   None,  None,  None,  None,  None,  ""],
        ["PT-LIN-01",      "ElementType",  "ET-LIN-01",      2, "ET-PROFILE-01",          None, 1,     None,   None,  None,  None,  None,  None,  ""],
        ["PT-LIN-01",      "ElementType",  "ET-LIN-01",      3, "ET-DIFF-01",             None, 1,     None,   None,  None,  None,  None,  None,  ""],
        ["PT-LIN-01",      "ElementType",  "ET-LIN-01",      4, "ET-CAP-01",              2,    None,  None,   None,  None,  "Y",   None,  None,  ""],

        # ─── PT-DL-EXT-01 — valid exterior downlight ──────────────────────────
        ["PT-DL-EXT-01",   "PositionType", "PT-DL-EXT-01",   1, "ET-DL-01",               None, None,  None,   None,  "Y",   None,  None,  None,  ""],
        ["PT-DL-EXT-01",   "PositionType", "PT-DL-EXT-01",   2, "ET-5Pin-Local-01",        1,   None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-EXT-01",   "ElementType",  "ET-DL-01",       1, "ET-CCL-D-700-1CH-01",    1,    None,  None,   None,  None,  "Y",   None,  None,  ""],

        # ─── PT-DL-TW-01 — valid twin spot ────────────────────────────────────
        ["PT-DL-TW-01",    "PositionType", "PT-DL-TW-01",    1, "ET-DL-01",               None, None,  None,   None,  "Y",   None,  None,  None,  ""],
        ["PT-DL-TW-01",    "PositionType", "PT-DL-TW-01",    2, "ET-5Pin-Local-01",        2,   None,  None,   None,  None,  "Y",   None,  None,  ""],
        ["PT-DL-TW-01",    "ElementType",  "ET-DL-01",       1, "ET-CCL-D-700-1CH-01",    2,    None,  None,   None,  None,  "Y",   None,  None,  ""],
    ]
    # fmt: on
    _write_sheet(wb, "Form", rs_headers, rs_rows)

    return _save(wb, "sample_rs.xlsx")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Writing fixture files to: {FIXTURE_DIR}")
    create_sample_db()
    create_sample_ps()
    create_sample_rs()
    print("Done.")


if __name__ == "__main__":
    main()
