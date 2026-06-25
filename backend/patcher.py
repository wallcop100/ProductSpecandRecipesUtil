"""
patcher.py — In-place xlsx patching with automatic backup for Recipe Builder.

Public API::

    backup_file(filepath: str) -> str
    patch_ps(filepath: str, changes: list) -> dict
    patch_rs(filepath: str, changes: list) -> dict

Each patch function returns::

    {"success": True, "backup_path": str}
    {"success": False, "error": str}
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

def backup_file(filepath: str) -> str:
    """
    Create a timestamped backup of *filepath* in the same directory.

    The backup filename follows the pattern::

        <original_stem>.backup.YYYYMMDD_HHMMSS.xlsx

    Returns the absolute path to the backup file.
    Raises ``OSError`` on failure.
    """
    filepath = os.path.abspath(filepath)
    directory = os.path.dirname(filepath)
    base = os.path.basename(filepath)

    # Strip .xlsx extension, add backup suffix, re-add extension
    if base.lower().endswith(".xlsx"):
        stem = base[:-5]
    else:
        stem = base

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{stem}.backup.{timestamp}.xlsx"
    backup_path = os.path.join(directory, backup_name)

    shutil.copy2(filepath, backup_path)
    return backup_path


# ---------------------------------------------------------------------------
# Internal patch engine
# ---------------------------------------------------------------------------

def _apply_patch(filepath: str, sheet_name: str, changes: list) -> dict:
    """
    Open *filepath*, locate *sheet_name*, apply *changes*, and save.

    Each change dict must have::

        {
            "row_index":  int,   # 1-based row index in the sheet
            "column":     int,   # 1-based column index in the sheet
            "value":      any    # new cell value (None clears the cell)
        }

    Returns the standard result dict.
    """
    # --- validate inputs -------------------------------------------------------
    if not os.path.isfile(filepath):
        return {"success": False, "error": f"File not found: {filepath}"}

    if not changes:
        return {"success": False, "error": "No changes provided."}

    for i, change in enumerate(changes):
        for key in ("row_index", "column"):
            if key not in change:
                return {"success": False, "error": f"Change #{i} is missing required key '{key}'."}
            if not isinstance(change[key], int) or change[key] < 1:
                return {"success": False, "error": f"Change #{i} '{key}' must be a positive integer."}

    # --- backup ----------------------------------------------------------------
    try:
        backup_path = backup_file(filepath)
    except OSError as exc:
        return {"success": False, "error": f"Could not create backup: {exc}"}

    # --- patch -----------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as exc:
        return {"success": False, "error": f"Could not open workbook: {exc}"}

    if sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{sheet_name}' not found in {filepath}"}

    ws = wb[sheet_name]

    for change in changes:
        row_idx = change["row_index"]
        col_idx = change["column"]
        value = change.get("value")  # None is valid — it clears the cell
        ws.cell(row=row_idx, column=col_idx).value = value

    try:
        wb.save(filepath)
    except Exception as exc:
        return {"success": False, "error": f"Could not save workbook: {exc}"}

    return {"success": True, "backup_path": backup_path}


# ---------------------------------------------------------------------------
# Public patch functions
# ---------------------------------------------------------------------------

def patch_ps(filepath: str, changes: list) -> dict:
    """
    Apply cell-level changes to the "Form" sheet of a Product Spec xlsx.

    Parameters
    ----------
    filepath:
        Absolute path to the PS xlsx file.
    changes:
        List of ``{"row_index": int, "column": int, "value": any}`` dicts.
        Row and column indices are **1-based** (row 1 is the header row).

    Returns
    -------
    ``{"success": True, "backup_path": str}`` on success.
    ``{"success": False, "error": str}`` on failure.
    """
    return _apply_patch(filepath, "Form", changes)


def patch_rs(filepath: str, changes: list) -> dict:
    """
    Apply cell-level changes to the "Form" sheet of a Recipe Spec xlsx.

    Parameters
    ----------
    filepath:
        Absolute path to the RS xlsx file.
    changes:
        List of ``{"row_index": int, "column": int, "value": any}`` dicts.
        Row and column indices are **1-based** (row 1 is the header row).

    Returns
    -------
    ``{"success": True, "backup_path": str}`` on success.
    ``{"success": False, "error": str}`` on failure.
    """
    return _apply_patch(filepath, "Form", changes)
