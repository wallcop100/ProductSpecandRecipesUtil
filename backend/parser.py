"""
parser.py — Excel file parsers for Recipe Builder.

Parses the three Excel workbooks:
  - DesignDB ([DB])  → parse_db()
  - Product Spec     ([PS]) → parse_ps()
  - Recipe Spec      ([RS]) → parse_rs()

Also exposes detect_file_type() and detect_files() used by the Flask app.

All parsers use header-based column matching so they are robust to column
reordering. Each data row also carries '_row_num' (1-based Excel row index)
so the patcher can write back to the correct cell.
"""

import os
import openpyxl

# ---------------------------------------------------------------------------
# Column maps  {excel_header: output_field_name}
# ---------------------------------------------------------------------------

ET_COLUMN_MAP = {
    'Ref':         'ElementTypeRef',
    'Name':        'Name',
    'Description': 'Description',
    'ParentRef':   'Family',
    'IsCollection':'IsCollection',
    'IsDeleted':   'IsDeleted',
    'SortOrder':   'SortOrder',
}

PT_COLUMN_MAP = {
    'Ref':                    'PositionTypeRef',
    'Name':                   'Name',
    'Description':            'Description',
    'ParentRef':              'ParentRef',
    'IsCollection':           'IsCollection',
    'IsDeleted':              'IsDeleted',
    'DriverLocation':         'DriverLocation',
    'SecondaryPowerType':     'SecondaryPowerType',
    'ControlTypeRef':         'ControlTypeRef',
    'SecondaryPowerNodes_+ve':'SecondaryPowerNodes_+ve',
}

PS_COLUMN_MAP = {
    'EntityRef':            'ElementTypeRef',
    'Manufacturer':         'Manufacturer',
    'ProductCode':          'ProductCode',
    'ComponentDescription': 'ComponentDescription',
    'InternalNotesText':    'InternalNotesText',
    'IsTBC':                'IsTBC',
    'IsDeleted':            'IsDeleted',
    'IsPropertiesTBC':      'IsPropertiesTBC',
}

RS_COLUMN_MAP = {
    'ContextType':           'ContextType',
    'ContextRef':            'ContextRef',
    'RecipeIndex':           'RecipeIndex',
    'EntityRef':             'ElementTypeRef',
    'Sort Order':            'SortOrder',
    'Quantity':              'Quantity',
    'PackQuantity':          'PackQuantity',
    'IsDeleted':             'IsDeleted',
    'IsDesign':              'IsDesign',
    'IsContractItem':        'IsContractItem',
    'IsTRItem':              'IsTRItem',
    'Dim_QuantityMultiplier':'Dim_QuantityMultiplier',
    'IsInteger':             'IsInteger',
}

ET_FLAG_COLS = {'IsCollection', 'IsDeleted'}
PT_FLAG_COLS = {'IsCollection', 'IsDeleted'}
PS_FLAG_COLS = {'IsTBC', 'IsDeleted', 'IsPropertiesTBC'}
RS_FLAG_COLS = {'IsDesign', 'IsContractItem', 'IsTRItem', 'IsInteger', 'IsDeleted'}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_sheet_by_headers(ws, column_map, flag_cols=None):
    """
    Parse a worksheet by matching actual column headers to output field names.

    column_map : {excel_header: output_field_name}
    flag_cols  : set of output field names that must normalise to 'Y' or None

    Returns list of row dicts. Each dict includes '_row_num' (1-based Excel
    row index, header = row 1, first data row = row 2).
    Fully-blank rows are skipped.
    """
    if flag_cols is None:
        flag_cols = set()

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    actual_headers = [str(h).strip() if h is not None else '' for h in first_row]

    col_index = {}
    for excel_name, output_name in column_map.items():
        try:
            col_index[output_name] = actual_headers.index(excel_name)
        except ValueError:
            pass  # column absent in this file version

    rows = []
    for row_num, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {'_row_num': row_num}
        all_none = True
        for output_name, idx in col_index.items():
            raw = row_values[idx] if idx < len(row_values) else None

            if raw is None:
                value = None
            elif output_name in flag_cols:
                value = 'Y' if (isinstance(raw, str) and raw.strip().upper() == 'Y') else None
            elif isinstance(raw, str):
                value = raw.strip() or None
            else:
                value = raw

            row_dict[output_name] = value
            if value is not None:
                all_none = False

        if not all_none:
            rows.append(row_dict)

    return rows


# ---------------------------------------------------------------------------
# Public parsers
# ---------------------------------------------------------------------------

def _collection_refs(raw_rows, ref_field, parent_field):
    """
    A "true collection" is any Ref that appears as another row's ParentRef —
    i.e. a category/grouping node with children. Leaf refs are real entities
    (including app wrapper ETs, which may carry IsCollection='Y' but have no
    children in the DB hierarchy, so they are NOT filtered out).
    """
    parents = set()
    for r in raw_rows:
        pref = r.get(parent_field)
        if pref:
            parents.add(pref)
    return parents


def parse_db(filepath):
    """
    Parse a DesignDB xlsx workbook.

    Returns::

        {
            "element_types":  [{ElementTypeRef, Name, Description, Family, ...}, ...],
            "position_types": [{PositionTypeRef, Name, Description, ...}, ...]
        }

    A ref used as another row's ParentRef is a true collection and is filtered
    out; deleted rows (IsDeleted='Y') are filtered too. ElementTypes keep
    _row_num (the DB ElementTypes sheet is writable via the catalogue export);
    PositionTypes stay read-only so their _row_num is stripped.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if 'ElementTypes' not in wb.sheetnames:
        raise ValueError(f"Sheet 'ElementTypes' not found in {filepath}")
    if 'PositionTypes' not in wb.sheetnames:
        raise ValueError(f"Sheet 'PositionTypes' not found in {filepath}")

    raw_ets = _parse_sheet_by_headers(wb['ElementTypes'], ET_COLUMN_MAP, ET_FLAG_COLS)
    et_collections = _collection_refs(raw_ets, 'ElementTypeRef', 'Family')
    element_types = []
    for r in raw_ets:
        ref = r.get('ElementTypeRef')
        if not ref:
            continue
        if r.get('IsDeleted') == 'Y':
            continue
        if ref in et_collections:
            continue
        r.pop('IsDeleted', None)
        # _row_num kept — ElementTypes is writable
        element_types.append(r)

    raw_pts = _parse_sheet_by_headers(wb['PositionTypes'], PT_COLUMN_MAP, PT_FLAG_COLS)
    pt_collections = _collection_refs(raw_pts, 'PositionTypeRef', 'ParentRef')
    position_types = []
    for r in raw_pts:
        ref = r.get('PositionTypeRef')
        if not ref:
            continue
        if r.get('IsDeleted') == 'Y':
            continue
        if ref in pt_collections:
            continue
        r.pop('IsCollection', None)
        r.pop('IsDeleted', None)
        r.pop('_row_num', None)
        position_types.append(r)

    return {
        'element_types': element_types,
        'position_types': position_types,
    }


def parse_ps(filepath):
    """
    Parse a Product Spec xlsx workbook.

    Returns a list of row dicts (including '_row_num') for rows that have
    an ElementTypeRef. '_row_num' is kept so the patcher can write back.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if 'Form' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Form' not found in {filepath}")

    rows = _parse_sheet_by_headers(wb['Form'], PS_COLUMN_MAP, PS_FLAG_COLS)
    return [r for r in rows if r.get('ElementTypeRef')]


def parse_rs(filepath):
    """
    Parse a Recipe Spec xlsx workbook.

    Performs a two-pass to add PositionTypeRef to every row:
      - PositionType-level rows: PositionTypeRef = ContextRef
      - ElementType-level rows:  PositionTypeRef = the position that uses
        this ElementType as its IsDesign element

    If an ElementType is used as design by multiple positions, its internal
    recipe rows are duplicated (one copy per position). Each copy retains
    the original '_row_num' so edits still map back to the correct Excel row.

    Orphan ElementType rows (no position uses their ContextRef as design)
    are dropped since they cannot be displayed in any position's view.

    Returns list of row dicts including 'PositionTypeRef' and '_row_num'.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if 'Form' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Form' not found in {filepath}")

    raw_rows = _parse_sheet_by_headers(wb['Form'], RS_COLUMN_MAP, RS_FLAG_COLS)

    # Pass 1: build {ET_ref: [position_refs]} from IsDesign=Y PositionType rows
    et_to_positions = {}
    for row in raw_rows:
        if (row.get('ContextType') == 'PositionType' and
                row.get('IsDesign') == 'Y' and
                row.get('ElementTypeRef') and row.get('ContextRef')):
            et_to_positions.setdefault(row['ElementTypeRef'], []).append(row['ContextRef'])

    # Pass 2: assign PositionTypeRef
    result = []
    for row in raw_rows:
        if row.get('ContextType') == 'PositionType':
            row['PositionTypeRef'] = row.get('ContextRef')
            result.append(row)
        elif row.get('ContextType') == 'ElementType':
            et_ref = row.get('ContextRef')
            pos_refs = et_to_positions.get(et_ref, [])
            for pos_ref in pos_refs:
                new_row = dict(row)
                new_row['PositionTypeRef'] = pos_ref
                result.append(new_row)

    return result


# ---------------------------------------------------------------------------
# File detection helpers (used by app.py /detect-files endpoint)
# ---------------------------------------------------------------------------

def detect_file_type(filepath):
    """
    Inspect a single xlsx file and return "db", "ps", "rs", or None.

    Detection rules:
      - "db" → has sheets "ElementTypes" AND "PositionTypes"
      - "ps" → has sheet "Form" AND row-1 contains header "ProductCode"
      - "rs" → has sheet "Form" AND row-1 contains header "ContextRef"
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None

    sheet_names = wb.sheetnames

    if 'ElementTypes' in sheet_names and 'PositionTypes' in sheet_names:
        wb.close()
        return 'db'

    if 'Form' in sheet_names:
        ws = wb['Form']
        first_row = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            first_row = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
            break
        wb.close()

        if 'ProductCode' in first_row:
            return 'ps'
        if 'ContextRef' in first_row:
            return 'rs'
        return None

    try:
        wb.close()
    except Exception:
        pass

    return None


def detect_files(folder):
    """
    Scan *folder* for .xlsx files and attempt to classify each as DB/PS/RS.

    Returns::

        {
            "db": "filename.xlsx" or None,
            "ps": "filename.xlsx" or None,
            "rs": "filename.xlsx" or None,
            "all_xlsx": ["file1.xlsx", ...]
        }
    """
    result = {'db': None, 'ps': None, 'rs': None, 'all_xlsx': []}

    if not os.path.isdir(folder):
        return result

    xlsx_files = [
        f for f in os.listdir(folder)
        if f.lower().endswith('.xlsx') and os.path.isfile(os.path.join(folder, f))
    ]
    result['all_xlsx'] = sorted(xlsx_files)

    for filename in xlsx_files:
        filepath = os.path.join(folder, filename)
        file_type = detect_file_type(filepath)
        if file_type and result[file_type] is None:
            result[file_type] = filename

    return result
