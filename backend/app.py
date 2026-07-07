"""
app.py — Flask backend for Recipe Builder.

Routes
------
GET  /ping
GET  /detect-files?folder=<path>
POST /import
POST /validate
POST /patch
"""

from __future__ import annotations

import os
import sys
import tempfile

from flask import Flask, jsonify, request

_BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
if _BACKEND_DIR not in sys.path:
    sys.path.insert(0, _BACKEND_DIR)

from parser import detect_files, parse_db, parse_ps, parse_rs  # noqa: E402
from validator import validate as _validate  # noqa: E402

import openpyxl  # noqa: E402

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = Flask(__name__)

from flask_cors import CORS  # noqa: E402
CORS(app)

# ---------------------------------------------------------------------------
# Field → Excel column name maps (used by the row-level patcher)
# ---------------------------------------------------------------------------

RS_FIELD_TO_EXCEL = {
    'ContextType':            'ContextType',
    'ContextRef':             'ContextRef',
    'RecipeIndex':            'RecipeIndex',
    'ElementTypeRef':         'EntityRef',
    'SortOrder':              'Sort Order',
    'Quantity':               'Quantity',
    'PackQuantity':           'PackQuantity',
    'IsDeleted':              'IsDeleted',
    'IsDesign':               'IsDesign',
    'IsContractItem':         'IsContractItem',
    'IsTRItem':               'IsTRItem',
    'Dim_QuantityMultiplier': 'Dim_QuantityMultiplier',
    'IsInteger':              'IsInteger',
}

PS_FIELD_TO_EXCEL = {
    'ElementTypeRef':      'EntityRef',
    'Manufacturer':        'Manufacturer',
    'ProductCode':         'ProductCode',
    'ComponentDescription':'ComponentDescription',
    'InternalNotesText':   'InternalNotesText',
    'IsTBC':               'IsTBC',
    'IsDeleted':           'IsDeleted',
    'IsPropertiesTBC':     'IsPropertiesTBC',
}

# DesignDB ElementTypes sheet — writable catalogue (EXPORT_PLAN §4).
# Identity + classification only; physical/electrical columns are left to the
# upstream pipeline and are never written by the app.
DB_FIELD_TO_EXCEL = {
    'ElementTypeRef': 'Ref',
    'Name':           'Name',
    'Description':    'Description',
    'Family':         'ParentRef',
    'IsCollection':   'IsCollection',
    'IsDeleted':      'IsDeleted',
    'SortOrder':      'SortOrder',
}


# ---------------------------------------------------------------------------
# Row-level patch engine
# ---------------------------------------------------------------------------

def _norm(value):
    """Normalise a cell/before value for comparison: blank strings → None."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def _values_differ(a, b):
    """Loose comparison: None == blank, 1 == 1.0, '1' == 1."""
    a, b = _norm(a), _norm(b)
    if a == b:
        return False
    # Numeric-vs-string tolerance ('1' on one side, 1 on the other)
    try:
        return float(a) != float(b)
    except (TypeError, ValueError):
        return True


def _apply_row_level_patch(filepath, sheet_name, field_to_excel, changes, key_column='EntityRef'):
    """
    Apply row-level changes to the 'Form' sheet of an xlsx file.

    Two change formats are supported:

    RS format:
        [{"_id": str, "action": "upsert"|"delete",
          "row": {..., "_row_num": int|None},
          "changedFields": {field: value} | None,   # field-level patch when present
          "before": {field: originalValue} | None}, ...]

    PS format:
        [{"elementTypeRef": str, "updates": {field: value},
          "before": {field: originalValue} | None, "_isNew": bool}, ...]

    Behaviour (see EXPORT_PLAN.md):
    - Staleness check: every change carrying 'before' has those values compared
      against the live cells. ANY mismatch blocks the entire file — nothing is
      written, no backup made, and the conflicts are returned.
    - Appends write EntityType='ElementType' plus mapped fields, and their
      assigned row numbers are returned in 'assignments' for reconciliation.
    - RS appends are guarded by a natural key (ContextType, ContextRef,
      RecipeIndex, EntityRef): if a matching row already exists on disk the
      append becomes an update at that row (never duplicates).
    - Deletes are tombstones only: IsDeleted='Y'. Exports are gospel.
    """
    if not os.path.isfile(filepath):
        return {'success': False, 'error': f'File not found: {filepath}'}

    if not changes:
        return {'success': False, 'error': 'No changes provided.'}

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as exc:
        return {'success': False, 'error': f'Could not open workbook: {exc}'}

    if sheet_name not in wb.sheetnames:
        return {'success': False, 'error': f"Sheet '{sheet_name}' not found"}

    ws = wb[sheet_name]

    # Read headers → 1-based column index map
    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val is not None:
            header_to_col[str(val).strip()] = c

    is_ps_format = changes and 'elementTypeRef' in changes[0]
    assignments = {}
    conflicts = []

    def check_before(key, row_num, before):
        """Compare a change's 'before' snapshot against live disk cells."""
        for field_name, original in (before or {}).items():
            excel_col = field_to_excel.get(field_name)
            if not excel_col or excel_col not in header_to_col:
                continue
            disk_value = ws.cell(row=row_num, column=header_to_col[excel_col]).value
            if _values_differ(original, disk_value):
                conflicts.append({
                    'key': key,
                    'field': field_name,
                    'column': excel_col,
                    'rowNum': row_num,
                    'diskValue': disk_value,
                    'baseValue': original,
                })

    if is_ps_format:
        entity_ref_col = header_to_col.get(key_column)
        if not entity_ref_col:
            return {'success': False, 'error': f"Column '{key_column}' not found in {sheet_name} sheet"}
        entity_type_col = header_to_col.get('EntityType')

        ref_to_row = {}
        for rn in range(2, ws.max_row + 1):
            val = ws.cell(row=rn, column=entity_ref_col).value
            if val:
                ref_to_row[str(val).strip()] = rn

        # Pass 1: staleness check — block the whole file on any conflict
        for change in changes:
            et_ref = change.get('elementTypeRef')
            row_num = ref_to_row.get(et_ref)
            if row_num:
                check_before(et_ref, row_num, change.get('before'))
        if conflicts:
            return {'success': False, 'conflict': True, 'conflicts': conflicts}

        # Pass 2: write
        for change in changes:
            et_ref = change.get('elementTypeRef')
            updates = change.get('updates') or {}
            is_new = change.get('_isNew', False)
            row_num = ref_to_row.get(et_ref)

            if not row_num:
                if not is_new:
                    continue
                # Append new row: EntityRef + EntityType, then non-null update fields
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=entity_ref_col).value = et_ref
                if entity_type_col:
                    ws.cell(row=row_num, column=entity_type_col).value = 'ElementType'
                ref_to_row[et_ref] = row_num
                assignments[et_ref] = row_num

            for field_name, value in updates.items():
                if value is None:
                    continue
                excel_col = field_to_excel.get(field_name)
                if excel_col and excel_col in header_to_col:
                    ws.cell(row=row_num, column=header_to_col[excel_col]).value = value

    else:
        # RS: deduplicate by _id (last write wins, delete trumps)
        latest = {}
        for change in changes:
            cid = change.get('_id')
            if not cid:
                continue
            action = change.get('action', 'upsert')
            if cid not in latest:
                latest[cid] = dict(change, action=action)
            else:
                if action == 'delete':
                    latest[cid]['action'] = 'delete'
                    if not latest[cid].get('row'):
                        latest[cid]['row'] = change.get('row') or {}
                else:
                    latest[cid] = dict(change, action=action)

        isdeleted_col = header_to_col.get('IsDeleted')
        entity_type_col = header_to_col.get('EntityType')

        # Natural-key index of existing rows: (ContextType, ContextRef,
        # RecipeIndex, EntityRef) → row_num. Guards appends against duplication
        # when a reconciliation payload was lost (crash mid-export).
        def nat_key_cols():
            cols = {}
            for h in ('ContextType', 'ContextRef', 'RecipeIndex', 'EntityRef'):
                cols[h] = header_to_col.get(h)
            return cols

        nk_cols = nat_key_cols()
        nat_index = {}
        if all(nk_cols.values()):
            for rn in range(2, ws.max_row + 1):
                key = tuple(
                    _norm(ws.cell(row=rn, column=nk_cols[h]).value)
                    for h in ('ContextType', 'ContextRef', 'RecipeIndex', 'EntityRef')
                )
                if any(v is not None for v in key):
                    nat_index.setdefault(key, rn)

        def natural_key(row):
            return (
                _norm(row.get('ContextType')),
                _norm(row.get('ContextRef')),
                _norm(row.get('RecipeIndex')),
                _norm(row.get('ElementTypeRef')),
            )

        # Pass 1: staleness check — block the whole file on any conflict
        for cid, entry in latest.items():
            row = entry.get('row') or {}
            row_num = row.get('_row_num')
            if row_num:
                check_before(cid, row_num, entry.get('before'))
        if conflicts:
            return {'success': False, 'conflict': True, 'conflicts': conflicts}

        # Pass 2: write
        for cid, entry in latest.items():
            action = entry['action']
            row = entry.get('row') or {}
            changed = entry.get('changedFields')
            row_num = row.get('_row_num')

            if action == 'delete':
                # Tombstone only. For never-reconciled rows, try the natural key
                # (the row may exist on disk from a lost reconciliation).
                if not row_num:
                    row_num = nat_index.get(natural_key(row))
                if row_num and isdeleted_col:
                    ws.cell(row=row_num, column=isdeleted_col).value = 'Y'

            elif action == 'upsert':
                is_append = not row_num
                if is_append:
                    # Natural-key guard: update in place if the row already exists
                    existing = nat_index.get(natural_key(row))
                    if existing:
                        row_num = existing
                        is_append = False
                    else:
                        row_num = ws.max_row + 1
                        if entity_type_col:
                            ws.cell(row=row_num, column=entity_type_col).value = 'ElementType'
                        nat_index[natural_key(row)] = row_num
                    assignments[cid] = row_num

                if changed and not is_append:
                    # Field-level patch: write only the changed fields
                    for field_name, value in changed.items():
                        excel_col = field_to_excel.get(field_name)
                        if excel_col and excel_col in header_to_col:
                            ws.cell(row=row_num, column=header_to_col[excel_col]).value = value
                else:
                    # Full-row write (appends, and legacy whole-row changes)
                    for field_name, excel_col_name in field_to_excel.items():
                        if excel_col_name not in header_to_col:
                            continue
                        if field_name not in row:
                            continue
                        ws.cell(
                            row=row_num,
                            column=header_to_col[excel_col_name],
                        ).value = row[field_name]

    # No per-write .backup file. Atomic save (below) already guarantees the
    # original is never corrupted by a failed write, and user-taken snapshots are
    # the rollback mechanism — so we keep a single set of .xlsx files instead of
    # accumulating a timestamped backup on every export. (backup_file is retained
    # in patcher.py for snapshots / explicit use.)
    backup_path = None

    # Atomic save: write to a temp file in the same directory, then replace the
    # original in one move. If openpyxl fails to serialise the workbook (e.g. it
    # can't round-trip some embedded content), the original file is left
    # completely untouched instead of being half-overwritten and corrupted.
    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix='.xlsx', dir=os.path.dirname(filepath) or '.'
    )
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
    except Exception as exc:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        return {
            'success': False,
            'error': f'Could not save workbook (your file was left untouched): {exc}',
        }

    try:
        os.replace(tmp_path, filepath)
    except OSError as exc:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        return {'success': False, 'error': f'Could not replace workbook: {exc}'}

    return {'success': True, 'backup_path': backup_path, 'assignments': assignments}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _error(message: str, status: int = 400):
    return jsonify({'error': message}), status


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok'})


@app.route('/detect-files', methods=['GET'])
def detect_files_route():
    folder = request.args.get('folder', '').strip()
    if not folder:
        return _error("'folder' query parameter is required.")
    result = detect_files(folder)
    return jsonify(result)


@app.route('/import', methods=['POST'])
def import_files():
    body = request.get_json(force=True, silent=True)
    if not body:
        return _error('Request body must be JSON.')

    db_path = body.get('db', '').strip()
    ps_path = body.get('ps', '').strip()
    rs_path = body.get('rs', '').strip()

    missing = [name for name, path in [('db', db_path), ('ps', ps_path), ('rs', rs_path)] if not path]
    if missing:
        return _error(f"Missing file path(s): {', '.join(missing)}")

    errors = []
    for label, path in [('db', db_path), ('ps', ps_path), ('rs', rs_path)]:
        if not os.path.isfile(path):
            errors.append(f"{label}: file not found at '{path}'")
    if errors:
        return _error('; '.join(errors))

    try:
        db_data = parse_db(db_path)
    except Exception as exc:
        return _error(f'Failed to parse DB file: {exc}')

    try:
        ps_rows = parse_ps(ps_path)
    except Exception as exc:
        return _error(f'Failed to parse PS file: {exc}')

    try:
        rs_rows = parse_rs(rs_path)
    except Exception as exc:
        return _error(f'Failed to parse RS file: {exc}')

    return jsonify({'db': db_data, 'ps': ps_rows, 'rs': rs_rows})


@app.route('/validate', methods=['POST'])
def validate_route():
    body = request.get_json(force=True, silent=True)
    if not body:
        return _error('Request body must be JSON.')

    db_data = body.get('db_data')
    ps_rows = body.get('ps_rows')
    rs_rows = body.get('rs_rows')

    if db_data is None or ps_rows is None or rs_rows is None:
        return _error("Body must include 'db_data', 'ps_rows', and 'rs_rows'.")

    issues = _validate(db_data, ps_rows, rs_rows)
    return jsonify({'issues': issues})


@app.route('/patch', methods=['POST'])
def patch_route():
    """
    Apply changes to a PS or RS xlsx file.

    Body (JSON)
    -----------
    {
        "target":   "ps" | "rs",
        "filepath": "/abs/path/to/file.xlsx",
        "changes":  [...]
    }

    RS changes: [{_id, action, row: {..., _row_num?}}]
    PS changes: [{elementTypeRef, updates: {field: value}}]
    """
    body = request.get_json(force=True, silent=True)
    if not body:
        return _error('Request body must be JSON.')

    target = body.get('target', '').strip().lower()
    filepath = body.get('filepath', '').strip()
    changes = body.get('changes')

    if target not in ('ps', 'rs', 'db'):
        return _error("'target' must be 'ps', 'rs', or 'db'.")
    if not filepath:
        return _error("'filepath' is required.")
    if not isinstance(changes, list) or not changes:
        return _error("'changes' must be a non-empty array.")

    if target == 'ps':
        result = _apply_row_level_patch(filepath, 'Form', PS_FIELD_TO_EXCEL, changes, key_column='EntityRef')
    elif target == 'db':
        # Writable DesignDB catalogue: patches the ElementTypes sheet, keyed on Ref.
        result = _apply_row_level_patch(filepath, 'ElementTypes', DB_FIELD_TO_EXCEL, changes, key_column='Ref')
    else:
        result = _apply_row_level_patch(filepath, 'Form', RS_FIELD_TO_EXCEL, changes)

    if result['success']:
        return jsonify(result)
    if result.get('conflict'):
        return jsonify(result), 409
    return jsonify(result), 422


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5001, debug=False)
